from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import re
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter, range_boundaries
import pandas as pd

from .config import FollowupError


PUNCT_RE = re.compile(r"[\W_]+", flags=re.UNICODE)


@dataclass
class DetectionResult:
    mapping: dict[str, str]
    notes: list[str]


def normalize_header(value: object) -> str:
    if value is None:
        return ""
    return " ".join(str(value).strip().lower().split())


def normalize_customer(value: object) -> str:
    if pd.isna(value):
        return ""
    s = str(value).upper()
    s = PUNCT_RE.sub("", s)
    return s


def parse_truthy(value: object) -> bool:
    if pd.isna(value):
        return False
    token = str(value).strip().upper()
    return token in {"TRUE", "1", "YES", "Y", "T"}


def parse_money(value: object) -> float | None:
    if pd.isna(value):
        return None
    s = str(value).strip().replace(",", "").replace("$", "")
    try:
        return float(s)
    except ValueError:
        return None


def read_excel(path: Path, sheet_name: str | None = None) -> pd.DataFrame:
    return pd.read_excel(path, sheet_name=sheet_name or 0, dtype=object)


def _find_header(headers: list[str], synonyms: Iterable[str], contains: str | None = None) -> str | None:
    normalized = {normalize_header(h): h for h in headers}
    for cand in synonyms:
        n = normalize_header(cand)
        if n in normalized:
            return normalized[n]
    if contains:
        for h in headers:
            if contains in normalize_header(h):
                return h
    return None


def detect_columns(
    df: pd.DataFrame,
    synonyms: dict[str, list[str]],
    required_fields: set[str],
    overrides: dict[str, str] | None = None,
    contains_rules: dict[str, str] | None = None,
) -> DetectionResult:
    overrides = overrides or {}
    contains_rules = contains_rules or {}
    headers = [str(c) for c in df.columns]
    mapping: dict[str, str] = {}
    notes: list[str] = []

    for field, col in overrides.items():
        if col in headers:
            mapping[field] = col
            notes.append(f"override: {field} -> {col}")

    for field, syns in synonyms.items():
        if field in mapping:
            continue
        found = _find_header(headers, syns, contains_rules.get(field))
        if found:
            mapping[field] = found

    missing = [f for f in required_fields if f not in mapping]
    if missing:
        detail = []
        for field in missing:
            tried = synonyms.get(field, [])
            detail.append(f"- {field}: tried {tried}")
        raise FollowupError(
            "Required field detection failed.\n"
            + "Missing fields:\n"
            + "\n".join(f"- {m}" for m in missing)
            + "\nExisting columns:\n"
            + "\n".join(f"- {h}" for h in headers)
            + "\nSynonyms tried:\n"
            + "\n".join(detail)
            + "\nUse --column-map mapping.json with {'quotes': {...}, 'orders': {...}} to override."
        )

    return DetectionResult(mapping=mapping, notes=notes)


def safe_excel_value(value: object) -> object:
    if isinstance(value, str) and value[:1] in {"=", "+", "-", "@"}:
        return "'" + value
    return value


def _find_header_row_and_columns(sheet, columns: list[str], scan_rows: int = 80) -> tuple[int, dict[str, int]]:
    targets = {normalize_header(c): c for c in columns}
    best_row = 1
    best_map: dict[str, int] = {}

    for r in range(1, min(scan_rows, sheet.max_row) + 1):
        found: dict[str, int] = {}
        for c in range(1, max(sheet.max_column, len(columns)) + 1):
            cell_value = sheet.cell(row=r, column=c).value
            n = normalize_header(cell_value)
            if n in targets and targets[n] not in found:
                found[targets[n]] = c
        if len(found) > len(best_map):
            best_row = r
            best_map = found
        if len(found) == len(columns):
            return r, found

    return best_row, best_map


def _find_matching_table(sheet, cols: list[str]):
    if not sheet.tables:
        return None

    wanted = {normalize_header(c): c for c in cols}
    for table in sheet.tables.values():
        min_col, min_row, max_col, _ = range_boundaries(table.ref)
        headers: dict[str, int] = {}
        for cidx in range(min_col, max_col + 1):
            label = sheet.cell(row=min_row, column=cidx).value
            n = normalize_header(label)
            if n in wanted:
                headers[wanted[n]] = cidx
        if all(c in headers for c in cols):
            return table, min_row, headers
    return None


def _write_to_existing_table(sheet, df: pd.DataFrame, table, header_row: int, positions: dict[str, int]) -> None:
    min_col, _, max_col, max_row = range_boundaries(table.ref)
    data_start = header_row + 1

    for r in range(data_start, max_row + 1):
        for c in range(min_col, max_col + 1):
            sheet.cell(row=r, column=c).value = None

    for ridx, row in enumerate(df.itertuples(index=False, name=None), start=data_start):
        for col, value in zip(df.columns, row):
            sheet.cell(row=ridx, column=positions[str(col)]).value = safe_excel_value(value)

    new_last_row = data_start + max(len(df), 1) - 1
    if len(df) == 0:
        for c in range(min_col, max_col + 1):
            sheet.cell(row=data_start, column=c).value = None

    table.ref = f"{get_column_letter(min_col)}{header_row}:{get_column_letter(max_col)}{new_last_row}"


def _write_dataframe_to_sheet(sheet, df: pd.DataFrame) -> None:
    cols = [str(c) for c in df.columns]
    table_match = _find_matching_table(sheet, cols)
    if table_match is not None:
        table, header_row, positions = table_match
        _write_to_existing_table(sheet, df, table, header_row, positions)
        return

    header_row, existing_positions = _find_header_row_and_columns(sheet, cols)

    positions = existing_positions.copy()
    next_col = (max(existing_positions.values()) + 1) if existing_positions else 1
    for col in cols:
        if col not in positions:
            positions[col] = next_col
            next_col += 1

    data_start = header_row + 1

    for col in cols:
        cidx = positions[col]
        sheet.cell(row=header_row, column=cidx).value = col
        for r in range(data_start, sheet.max_row + 1):
            sheet.cell(row=r, column=cidx).value = None

    for row_offset, row in enumerate(df.itertuples(index=False, name=None), start=0):
        ridx = data_start + row_offset
        for col, value in zip(cols, row):
            cidx = positions[col]
            sheet.cell(row=ridx, column=cidx).value = safe_excel_value(value)


def write_output(path: Path, sheets: dict[str, pd.DataFrame], template_path: Path | None = None) -> None:
    if template_path:
        wb = load_workbook(template_path, keep_links=False)
        wb._external_links = []
    else:
        wb = Workbook()
        default = wb.active
        wb.remove(default)

    for sheet_name, df in sheets.items():
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(title=sheet_name)
        _write_dataframe_to_sheet(ws, df)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
