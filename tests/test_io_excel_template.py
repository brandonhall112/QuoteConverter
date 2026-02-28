from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
import pandas as pd

from followup_quotes.app import generate_followup_workbook, resolve_template_path
from followup_quotes.config import RunConfig
from followup_quotes.io_excel import write_output


def test_write_output_preserves_template_formula_sheet_and_checkbox_column(tmp_path: Path):
    template = tmp_path / "template.xlsx"
    out = tmp_path / "out.xlsx"

    wb = Workbook()
    ws_followup = wb.active
    ws_followup.title = "Follow-Up"
    ws_followup["A1"] = "Quote"
    ws_followup["B1"] = "Customer"
    ws_followup["C1"] = "Won by Follow Up?"
    ws_followup["C2"] = True

    ws_summary = wb.create_sheet("Summary")
    ws_summary["A1"] = "Count"
    ws_summary["B1"] = "=COUNTA('Follow-Up'!A:A)-1"
    ws_summary["C1"] = "FollowupOneCount"
    ws_summary["D1"] = "=COUNTIF('Follow-Up'!C:C,TRUE)"
    wb.save(template)

    sheets = {
        "Follow-Up": pd.DataFrame(
            [
                {"Quote": "Q1", "Customer": "ACME", "Won by Follow Up?": False},
                {"Quote": "Q2", "Customer": "BETA", "Won by Follow Up?": False},
            ]
        ),
        "_Meta": pd.DataFrame([{"Metric": "x", "Value": "y"}]),
    }

    write_output(out, sheets, template)

    out_wb = load_workbook(out)
    assert out_wb["Follow-Up"]["A2"].value == "Q1"
    assert out_wb["Follow-Up"]["A3"].value == "Q2"
    assert out_wb["Follow-Up"]["C2"].value is False
    assert out_wb["Summary"]["B1"].value == "=COUNTA('Follow-Up'!A:A)-1"
    assert out_wb["Summary"]["D1"].value == "=COUNTIF('Follow-Up'!C:C,TRUE)"


def test_generate_followup_workbook_creates_per_rep_tabs(tmp_path: Path, monkeypatch):
    # Ensure this test is deterministic and not influenced by repo-level auto-detected templates.
    monkeypatch.chdir(tmp_path)

    quotes_path = tmp_path / "quotes.xlsx"
    orders_path = tmp_path / "orders.xlsx"
    out_path = tmp_path / "output.xlsx"

    pd.DataFrame(
        {
            "Quote #": ["Q1", "Q2"],
            "Customer": ["Acme", "Beta"],
            "Amount": [4000, 4100],
            "Date Quoted": ["2024-01-01", "2024-01-02"],
            "Entry Person Name": ["Reid Kincaid", "Eric Simpson"],
        }
    ).to_excel(quotes_path, index=False)

    pd.DataFrame(
        {
            "Order Number": [1],
            "Customer": ["ACME"],
            "Net Amount": [4000],
        }
    ).to_excel(orders_path, index=False)

    cfg = RunConfig(
        quotes_path=quotes_path,
        orders_path=orders_path,
        out_path=out_path,
        reps=["Reid Kincaid", "Eric Simpson"],
    )

    generate_followup_workbook(cfg)

    wb = load_workbook(out_path)
    assert "Follow-Up" in wb.sheetnames
    assert "Eric Simpson" in wb.sheetnames
    assert wb["Eric Simpson"]["A2"].value == "Q2"


def test_resolve_template_path_prefers_explicit(tmp_path: Path):
    explicit = tmp_path / "x.xlsx"
    explicit.write_bytes(b"dummy")
    assert resolve_template_path(explicit) == explicit


def test_write_output_updates_existing_table_ref(tmp_path: Path):
    template = tmp_path / "table_template.xlsx"
    out = tmp_path / "table_out.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Follow-Up"
    ws.append(["Quote", "Customer", "Quote Amount", "Date Quoted", "Entry Person Name", "Won by Follow Up?"])
    ws.append(["Q-old", "ACME", 10, "2024-01-01", "Reid", False])
    ws.append(["Q-old2", "ACME", 11, "2024-01-01", "Reid", False])

    table = Table(displayName="Table1", ref="A1:F3")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(table)
    wb.save(template)

    df = pd.DataFrame([
        {"Quote": "Q1", "Customer": "ACME", "Quote Amount": 100, "Date Quoted": "2024-02-01", "Entry Person Name": "Reid", "Won by Follow Up?": False},
        {"Quote": "Q2", "Customer": "BETA", "Quote Amount": 200, "Date Quoted": "2024-02-02", "Entry Person Name": "Eric", "Won by Follow Up?": True},
        {"Quote": "Q3", "Customer": "GAMMA", "Quote Amount": 300, "Date Quoted": "2024-02-03", "Entry Person Name": "Eric", "Won by Follow Up?": False},
    ])

    write_output(out, {"Follow-Up": df}, template)

    out_wb = load_workbook(out)
    out_ws = out_wb["Follow-Up"]
    assert out_ws.tables["Table1"].ref == "A1:F4"
    assert out_ws["A4"].value == "Q3"


def test_write_output_new_sheet_starts_at_column_a(tmp_path: Path):
    out = tmp_path / "plain.xlsx"
    write_output(
        out,
        {
            "Eric Simpson": pd.DataFrame([{"Quote": "Q2", "Customer": "BETA", "Won by Follow Up?": False}])
        },
        template_path=None,
    )

    wb = load_workbook(out)
    ws = wb["Eric Simpson"]
    assert ws["A1"].value == "Quote"
    assert ws["A2"].value == "Q2"


def test_resolve_template_path_prefers_parts_followup_template_in_assets(tmp_path: Path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    assets = tmp_path / "assets"
    assets.mkdir()
    preferred = assets / "Parts Follow Up Template.xlsx"
    preferred.write_bytes(b"dummy")

    assert resolve_template_path(None) == preferred
